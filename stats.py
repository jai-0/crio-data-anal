import openpyxl
from scrapper import *
import time
from datetime import date
import gspread
import sys
import os
import time
import schedule
from oauth2client.service_account import ServiceAccountCredentials
cwd=os.getcwd()
from keep_alive import keep_alive


# use creds to create a client to interact with the Google Drive API
scope = ['https://www.googleapis.com/auth/spreadsheets', "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
creds = ServiceAccountCredentials.from_json_keyfile_name(f'{cwd}/driveapiservice.json', scope)
client = gspread.authorize(creds)

# Find a workbook by name and open the first sheet
# Make sure you use the right name here.
sheetkey=os.getenv("sheet_key")
sheet = client.open_by_key(f"{sheetkey}")


today = date.today()
d1 = today.strftime("%d/%m/%Y")
d1=str(d1)

keep_alive()

def crawler(): 
    def getSpreadsheet():      
            url = 'https://docs.google.com/spreadsheets/d/1i1090-3GFWl6YVxF68j2JoRNwd2vc5xKSybu0QLjxt8/export?format=xlsx'
            r = requests.get(url, allow_redirects=True)
            open(f'{cwd}/bot of Social Media Stats.xlsx', 'wb').write(r.content)
    getSpreadsheet()
    time.sleep(30)
    wb = openpyxl.load_workbook(f'{cwd}/bot of Social Media Stats.xlsx')
    def getDay():
        intDay=today.weekday()
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        day=days[intDay]
        return day

    def getCell():
        worksh=wb.worksheets[0]
        for x in range(3,worksh.max_column):
            if worksh.cell(row=3, column=x).value == None and worksh.cell(row=3, column=x-1).value != None:
                return x
                break
    totalnum=0    
    x=getCell()
    day=getDay()            
    for ws in wb:
        for socials in range(3,10):
            data=""
            try:
                worksheet=sheet.worksheet(f"{ws.title}")
                link=str(ws.cell(row=socials, column=1).hyperlink.target)
                data= scrape(link)
                ws.cell(row=1,column=x).value=day
                worksheet.update_cell(1,x,f"{day}")
                ws.cell(row=2, column=x).value = str(d1)
                worksheet.update_cell(2,x,f"'{str(d1)}'")
                ws.cell(row=socials, column=x).value = data
                worksheet.update_cell(socials,x,f"{data}")
                wb.save(f'{cwd}/bot of Social Media Stats.xlsx')
                #print(f"{ws.cell(row=1,column=1).value}'s {ws.cell(row=socials, column=1).value} done ")
                         
                
            except:
                continue
        totalnum=totalnum+1
    os.remove(f'{cwd}/bot of Social Media Stats.xlsx')
schedule.every().day.at("16:30").do(crawler) 
while True:
    schedule.run_pending()
    